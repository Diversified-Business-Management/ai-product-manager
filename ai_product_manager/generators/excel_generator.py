"""
AI Product Manager - Excel Output Generator
Creates a scored and formatted Excel workbook from the scoring pipeline results.
"""
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


COLORS = {
    "header_bg": "065A82",
    "header_fg": "FFFFFF",
    "high": "059669",
    "medium": "D97706",
    "low": "DC2626",
    "alt_row": "F8FAFC",
    "white": "FFFFFF",
    "border": "E2E8F0",
}

def generate_scored_workbook(capabilities, survey_insights, summary_stats, output_path):
    wb = Workbook()
    _create_summary_sheet(wb, summary_stats, survey_insights)
    _create_rankings_sheet(wb, capabilities)
    _create_survey_sheet(wb, survey_insights)
    _create_config_sheet(wb)
    wb.save(output_path)

    # Recalculate formulas
    recalc_script = Path(__file__).parent.parent.parent / "mnt" / ".skills" / "skills" / "xlsx" / "scripts" / "recalc.py"
    if recalc_script.exists():
        try:
            subprocess.run(["python3", str(recalc_script), str(output_path)], timeout=30, capture_output=True)
        except Exception:
            pass
    return output_path


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color=COLORS["header_fg"], size=10, name="Arial")
        cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_data_rows(ws, start_row, end_row, max_col, priority_col=None):
    thin_border = Border(
        bottom=Side(style="thin", color=COLORS["border"])
    )
    for row in range(start_row, end_row + 1):
        bg = COLORS["alt_row"] if (row - start_row) % 2 == 0 else COLORS["white"]
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=10, name="Arial")
            cell.border = thin_border
            if col != 2:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        if priority_col:
            p_cell = ws.cell(row=row, column=priority_col)
            val = str(p_cell.value or "")
            if val == "HIGH":
                p_cell.font = Font(bold=True, color=COLORS["high"], size=10, name="Arial")
            elif val == "MEDIUM":
                p_cell.font = Font(bold=True, color=COLORS["medium"], size=10, name="Arial")
            elif val == "LOW":
                p_cell.font = Font(bold=True, color=COLORS["low"], size=10, name="Arial")


def _create_summary_sheet(wb, stats, insights):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = "065A82"

    ws["A1"] = "AI Product Manager — Scoring Dashboard"
    ws["A1"].font = Font(bold=True, size=16, name="Arial", color=COLORS["header_bg"])
    ws.merge_cells("A1:F1")

    labels = ["Total Capabilities", "High Priority", "Medium Priority", "Low Priority", "Avg Score", "Max Score", "Min Score", "Std Dev", "Survey Respondents"]
    values = [stats.get("total_capabilities", 0), stats.get("high_priority", 0), stats.get("medium_priority", 0),
              stats.get("low_priority", 0), stats.get("avg_score", 0), stats.get("max_score", 0),
              stats.get("min_score", 0), stats.get("std_dev", 0), insights.get("total_respondents", 0)]

    for i, (label, val) in enumerate(zip(labels, values)):
        row = i + 3
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10, name="Arial")
        ws.cell(row=row, column=2, value=val).font = Font(size=10, name="Arial")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15


def _create_rankings_sheet(wb, capabilities):
    ws = wb.create_sheet("Priority Rankings")
    ws.sheet_properties.tabColor = "059669"

    headers = ["Rank", "Capability", "Tier", "Category", "Priority", "Final Score",
               "Customer Impact", "Business Impact", "Cost to Implement", "Survey Demand", "Dev Quarter"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    for i, cap in enumerate(capabilities):
        row = i + 2
        ws.cell(row=row, column=1, value=cap["rank"])
        ws.cell(row=row, column=2, value=cap["name"])
        ws.cell(row=row, column=3, value=cap["tier"])
        ws.cell(row=row, column=4, value=cap["sheet"])
        ws.cell(row=row, column=5, value=cap["priority"])
        ws.cell(row=row, column=6, value=cap["scores"]["final"])
        ws.cell(row=row, column=7, value=cap["scores"]["customer_impact"])
        ws.cell(row=row, column=8, value=cap["scores"]["business_impact"])
        ws.cell(row=row, column=9, value=cap["scores"]["cost_to_implement"])
        ws.cell(row=row, column=10, value=cap.get("survey_demand", 0))
        ws.cell(row=row, column=11, value=cap.get("development_quarter", ""))

    _style_data_rows(ws, 2, len(capabilities) + 1, len(headers), priority_col=5)

    widths = [6, 45, 12, 15, 10, 12, 14, 14, 16, 14, 12]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    ws.auto_filter.ref = f"A1:K{len(capabilities) + 1}"
    ws.freeze_panes = "A2"


def _create_survey_sheet(wb, insights):
    ws = wb.create_sheet("Survey Analysis")
    ws.sheet_properties.tabColor = "D97706"

    ws["A1"] = "Customer Survey Analysis"
    ws["A1"].font = Font(bold=True, size=14, name="Arial", color=COLORS["header_bg"])

    ws["A3"] = "Capability"
    ws["B3"] = "Demand Count"
    ws["C3"] = "% of Respondents"
    ws["D3"] = "Requesting Companies"
    _style_header(ws, 3, 4)

    demand = insights.get("capability_demand", {})
    sorted_demand = sorted(demand.items(), key=lambda x: -x[1]["demand_count"])
    for i, (cap, data) in enumerate(sorted_demand):
        row = i + 4
        ws.cell(row=row, column=1, value=cap)
        ws.cell(row=row, column=2, value=data["demand_count"])
        ws.cell(row=row, column=3, value=data["demand_pct"] / 100)
        ws.cell(row=row, column=3).number_format = "0.0%"
        ws.cell(row=row, column=4, value=", ".join(data["requesting_companies"][:5]))

    _style_data_rows(ws, 4, len(sorted_demand) + 3, 4)
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 50


def _create_config_sheet(wb):
    ws = wb.create_sheet("Scoring Config")
    ws.sheet_properties.tabColor = "64748B"

    ws["A1"] = "Scoring Configuration"
    ws["A1"].font = Font(bold=True, size=14, name="Arial", color=COLORS["header_bg"])

    config_data = [
        ("", ""),
        ("DIMENSION WEIGHTS", ""),
        ("Customer Impact", 0.30),
        ("Business Impact", 0.50),
        ("Cost to Implement", 0.20),
        ("", ""),
        ("CUSTOMER IMPACT SUB-WEIGHTS", ""),
        ("Breadth of Customers", 0.50),
        ("Severity of Pain", 0.25),
        ("Level of Pain Mitigation", 0.25),
        ("", ""),
        ("BUSINESS IMPACT SUB-WEIGHTS", ""),
        ("New ACV", 0.30),
        ("Sales & Retention", 0.30),
        ("Strategic Value", 0.30),
        ("Competitive USP", 0.10),
        ("", ""),
        ("COST SUB-WEIGHTS", ""),
        ("Engineering Cost", 0.60),
        ("Ability to Execute", 0.25),
        ("COGS Impact", 0.15),
        ("", ""),
        ("PRIORITY THRESHOLDS", ""),
        ("High Priority", 7.0),
        ("Medium Priority", 5.0),
        ("Low Priority", 0.0),
    ]

    for i, (label, val) in enumerate(config_data):
        row = i + 3
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_b = ws.cell(row=row, column=2, value=val)
        if label.isupper() and label:
            cell_a.font = Font(bold=True, size=10, name="Arial", color=COLORS["header_bg"])
        else:
            cell_a.font = Font(size=10, name="Arial")
        if isinstance(val, float) and val < 1:
            cell_b.number_format = "0%"
        cell_b.font = Font(size=10, name="Arial", color="0000FF")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
