"""
AI Product Manager - Scoring Engine
Applies weighted prioritization to product capabilities using survey data and business scoring.
"""
import json
import pandas as pd
import numpy as np
from pathlib import Path
from collections import Counter, defaultdict


class ScoringEngine:
    def __init__(self, config_path=None):
        if config_path is None:
            config_path = Path(__file__).parent.parent / "config" / "scoring_config.json"
        with open(config_path) as f:
            self.config = json.load(f)
        self.weights = self.config["score_weights"]
        self.ai_adjustments = self.config["ai_adjustments"]
        self.tiers = self.config["monetization_tiers"]
        self.thresholds = self.config["scoring_scale"]["thresholds"]

    def compute_customer_impact(self, breadth, severity, mitigation):
        sw = self.weights["customer_impact"]["sub_weights"]
        return (breadth * sw["breadth_of_customers"] +
                severity * sw["severity_of_pain"] +
                mitigation * sw["level_of_pain_mitigation"])

    def compute_business_impact(self, new_acv, sales_retention, strategic_value, competitive_usp):
        sw = self.weights["business_impact"]["sub_weights"]
        return (new_acv * sw["new_acv"] +
                sales_retention * sw["sales_retention"] +
                strategic_value * sw["strategic_value"] +
                competitive_usp * sw["competitive_usp"])

    def compute_cost_to_implement(self, eng_cost, ability_to_execute, cogs_impact=None):
        sw = self.weights["cost_to_implement"]["sub_weights"]
        score = eng_cost * sw["engineering_cost"] + ability_to_execute * sw["ability_to_execute"]
        if cogs_impact is not None:
            score += cogs_impact * sw["cogs_impact"]
        return score

    def compute_total_score(self, customer_impact, business_impact, cost_to_implement):
        w = self.weights
        return (customer_impact * w["customer_impact"]["weight"] +
                business_impact * w["business_impact"]["weight"] +
                cost_to_implement * w["cost_to_implement"]["weight"])

    def apply_ai_adjustments(self, base_score, survey_demand_count=0, has_competitive_gap=False, quarters_out=0):
        adjusted = base_score
        if self.ai_adjustments["survey_demand_boost"]["enabled"] and survey_demand_count > 0:
            thresholds = self.ai_adjustments["survey_demand_boost"]["thresholds"]
            if survey_demand_count >= 7:
                adjusted += thresholds["7_plus_customers"]
            elif survey_demand_count >= 5:
                adjusted += thresholds["5_plus_customers"]
            elif survey_demand_count >= 3:
                adjusted += thresholds["3_plus_customers"]
        if self.ai_adjustments["competitive_gap_penalty"]["enabled"] and has_competitive_gap:
            adjusted += self.ai_adjustments["competitive_gap_penalty"]["penalty"]
        if self.ai_adjustments["time_decay"]["enabled"] and quarters_out > 0:
            decay = self.ai_adjustments["time_decay"]["decay_rate_per_quarter"] * quarters_out
            adjusted -= decay
        return max(0, min(10, adjusted))

    def classify_priority(self, score):
        if score >= self.thresholds["high_priority"]:
            return "HIGH"
        elif score >= self.thresholds["medium_priority"]:
            return "MEDIUM"
        return "LOW"

    def apply_tier_multiplier(self, score, tier):
        multiplier = self.tiers.get(tier, {}).get("multiplier", 1.0)
        return score * multiplier


class SurveyAnalyzer:
    """Analyzes customer survey responses to extract demand signals."""

    def __init__(self, survey_path):
        self.df = pd.read_csv(survey_path)
        self._clean_data()

    def _clean_data(self):
        # Skip header rows and test data
        valid_mask = self.df["Company"].notna() & ~self.df["Company"].isin([
            "Company", "Name", "NAME11", "COMPANY11", "CCCompany", "ERROL", ""
        ])
        self.df = self.df[valid_mask].reset_index(drop=True)

    def get_company_count(self):
        return len(self.df)

    def get_companies(self):
        return self.df["Company"].tolist()

    def analyze_capability_demand(self):
        """Map survey responses to capability demand counts."""
        capability_keywords = {
            "Catalog": ["catalog", "merchandising", "product"],
            "Bundles": ["bundle", "bundles"],
            "Pricing": ["pricing", "price", "waterfall"],
            "Promotions": ["promotion", "promo", "coupon", "offer"],
            "Storefront": ["storefront", "portal", "checkout", "cart"],
            "AI Optimization": ["ai powered", "ai ", "best offer", "optimization"],
            "Dynamic Pricing": ["dynamic pricing", "attribute based", "rules"],
            "Features & Entitlements": ["features", "entitlements"],
            "Journey Orchestration": ["journey", "orchestration", "campaign"],
            "Commerce Platform": ["commerce", "discover", "buy", "manage"],
        }
        demand = {}
        text_cols = [c for c in self.df.columns if self.df[c].dtype == object]
        for cap, keywords in capability_keywords.items():
            count = 0
            requesting_companies = []
            for _, row in self.df.iterrows():
                combined = " ".join(str(row[c]).lower() for c in text_cols if pd.notna(row[c]))
                if any(kw in combined for kw in keywords):
                    count += 1
                    requesting_companies.append(row.get("Company", "Unknown"))
            demand[cap] = {
                "demand_count": count,
                "requesting_companies": requesting_companies,
                "demand_pct": round(count / max(len(self.df), 1) * 100, 1)
            }
        return demand

    def analyze_business_models(self):
        col = "Business Model"
        if col not in self.df.columns:
            return {}
        models = Counter()
        for val in self.df[col].dropna():
            for m in str(val).split(","):
                m = m.strip()
                if m:
                    models[m] += 1
        return dict(models)

    def analyze_platforms(self):
        col = "Do you currently have a separate commerce platform? If so, please specify."
        if col not in self.df.columns:
            return {}
        platforms = Counter()
        for val in self.df[col].dropna():
            for p in str(val).split(","):
                p = p.strip()
                if p:
                    platforms[p] += 1
        return dict(platforms)

    def get_customer_insights(self):
        return {
            "total_respondents": self.get_company_count(),
            "companies": self.get_companies(),
            "capability_demand": self.analyze_capability_demand(),
            "business_models": self.analyze_business_models(),
            "existing_platforms": self.analyze_platforms(),
        }


class CapabilityScorer:
    """Scores capabilities from the Excel workbook and enriches with survey data."""

    def __init__(self, xlsx_path, survey_analyzer, scoring_engine):
        self.xlsx_path = xlsx_path
        self.survey = survey_analyzer
        self.engine = scoring_engine
        self.capabilities = []

    def load_and_score(self):
        """Load capability data from Excel and compute scores."""
        from openpyxl import load_workbook
        wb = load_workbook(self.xlsx_path, data_only=True)
        survey_demand = self.survey.analyze_capability_demand()

        # Process main Capability Scoring sheet
        self._process_sheet(wb["Capability Scoring"], survey_demand, "main")
        if "Dynamic Pricing" in wb.sheetnames:
            self._process_sheet(wb["Dynamic Pricing"], survey_demand, "pricing")
        if " Stack Ranked Custom Logic" in wb.sheetnames:
            self._process_custom_logic(wb[" Stack Ranked Custom Logic"], survey_demand)
        if "Capability Scoring Experiences" in wb.sheetnames:
            self._process_sheet(wb["Capability Scoring Experiences"], survey_demand, "experiences")

        self._rank_capabilities()
        return self.capabilities

    def _safe_float(self, val, default=0):
        try:
            return float(val) if val is not None else default
        except (ValueError, TypeError):
            return default

    def _process_sheet(self, ws, survey_demand, sheet_type):
        for row_idx in range(20, min(ws.max_row + 1, 100)):
            name = ws.cell(row=row_idx, column=1).value
            if not name or str(name).strip() == "":
                continue

            tier = str(ws.cell(row=row_idx, column=2).value or "Core")
            breadth = self._safe_float(ws.cell(row=row_idx, column=3).value)
            severity = self._safe_float(ws.cell(row=row_idx, column=5).value)
            mitigation = self._safe_float(ws.cell(row=row_idx, column=7).value)
            new_acv = self._safe_float(ws.cell(row=row_idx, column=10).value)
            sales_ret = self._safe_float(ws.cell(row=row_idx, column=12).value)
            strategic = self._safe_float(ws.cell(row=row_idx, column=14).value)
            competitive = self._safe_float(ws.cell(row=row_idx, column=16).value)
            eng_cost = self._safe_float(ws.cell(row=row_idx, column=19).value)
            ability = self._safe_float(ws.cell(row=row_idx, column=21).value)

            customer_impact = self.engine.compute_customer_impact(breadth, severity, mitigation)
            business_impact = self.engine.compute_business_impact(new_acv, sales_ret, strategic, competitive)
            cost = self.engine.compute_cost_to_implement(eng_cost, ability)
            base_score = self.engine.compute_total_score(customer_impact, business_impact, cost)

            # Find survey demand match
            demand_count = 0
            matched_cap = None
            for cap_key, demand_data in survey_demand.items():
                if any(kw in str(name).lower() for kw in cap_key.lower().split()):
                    demand_count = max(demand_count, demand_data["demand_count"])
                    matched_cap = cap_key

            ai_score = self.engine.apply_ai_adjustments(base_score, survey_demand_count=demand_count)
            final_score = self.engine.apply_tier_multiplier(ai_score, tier)
            priority = self.engine.classify_priority(final_score)

            self.capabilities.append({
                "name": str(name).strip(),
                "tier": tier,
                "sheet": sheet_type,
                "scores": {
                    "customer_impact": round(customer_impact, 2),
                    "business_impact": round(business_impact, 2),
                    "cost_to_implement": round(cost, 2),
                    "base_total": round(base_score, 2),
                    "ai_adjusted": round(ai_score, 2),
                    "final": round(final_score, 2),
                },
                "raw_inputs": {
                    "breadth": breadth, "severity": severity, "mitigation": mitigation,
                    "new_acv": new_acv, "sales_retention": sales_ret,
                    "strategic_value": strategic, "competitive_usp": competitive,
                    "engineering_cost": eng_cost, "ability_to_execute": ability,
                },
                "survey_demand": demand_count,
                "matched_survey_capability": matched_cap,
                "priority": priority,
                "rank": 0,
            })

    def _process_custom_logic(self, ws, survey_demand):
        for row_idx in range(20, min(ws.max_row + 1, 100)):
            name = ws.cell(row=row_idx, column=1).value
            if not name or str(name).strip() == "" or str(name).strip() == "Post-GA Continuation":
                continue

            tier = str(ws.cell(row=row_idx, column=4).value or "Ext Studio")
            quarter = str(ws.cell(row=row_idx, column=2).value or "TBD")
            notes = str(ws.cell(row=row_idx, column=3).value or "")

            breadth = self._safe_float(ws.cell(row=row_idx, column=5).value)
            severity = self._safe_float(ws.cell(row=row_idx, column=7).value)
            mitigation = self._safe_float(ws.cell(row=row_idx, column=9).value)
            new_acv = self._safe_float(ws.cell(row=row_idx, column=12).value)
            sales_ret = self._safe_float(ws.cell(row=row_idx, column=14).value)
            strategic = self._safe_float(ws.cell(row=row_idx, column=16).value)
            competitive = self._safe_float(ws.cell(row=row_idx, column=18).value)
            eng_cost = self._safe_float(ws.cell(row=row_idx, column=21).value)
            ability = self._safe_float(ws.cell(row=row_idx, column=23).value)
            cogs = self._safe_float(ws.cell(row=row_idx, column=25).value)

            customer_impact = self.engine.compute_customer_impact(breadth, severity, mitigation)
            business_impact = self.engine.compute_business_impact(new_acv, sales_ret, strategic, competitive)
            cost = self.engine.compute_cost_to_implement(eng_cost, ability, cogs)
            base_score = self.engine.compute_total_score(customer_impact, business_impact, cost)

            quarters_out = 0
            if "Q1 26" in quarter:
                quarters_out = 1
            elif "Q2 26" in quarter:
                quarters_out = 2
            elif "Q1-Q2" in quarter:
                quarters_out = 1

            ai_score = self.engine.apply_ai_adjustments(base_score, quarters_out=quarters_out)
            final_score = self.engine.apply_tier_multiplier(ai_score, tier)
            priority = self.engine.classify_priority(final_score)

            self.capabilities.append({
                "name": str(name).strip(),
                "tier": tier,
                "sheet": "custom_logic",
                "development_quarter": quarter,
                "notes": notes[:200],
                "scores": {
                    "customer_impact": round(customer_impact, 2),
                    "business_impact": round(business_impact, 2),
                    "cost_to_implement": round(cost, 2),
                    "base_total": round(base_score, 2),
                    "ai_adjusted": round(ai_score, 2),
                    "final": round(final_score, 2),
                },
                "survey_demand": 0,
                "priority": priority,
                "rank": 0,
            })

    def _rank_capabilities(self):
        # Remove duplicates (same name appearing across sheets) — keep highest score
        seen = {}
        for cap in self.capabilities:
            name = cap["name"]
            if name not in seen or cap["scores"]["final"] > seen[name]["scores"]["final"]:
                seen[name] = cap
        self.capabilities = list(seen.values())

        # Remove capabilities with zero/near-zero scores (empty rows) and invalid entries
        skip_names = {"#REF!", "Logic AI:", "Advanced Tier [TBD]", "Post-GA Continuation", "etc...."}
        self.capabilities = [
            c for c in self.capabilities
            if c["scores"]["final"] > 0.5 and c["name"] not in skip_names
        ]

        self.capabilities.sort(key=lambda c: c["scores"]["final"], reverse=True)
        for i, cap in enumerate(self.capabilities):
            cap["rank"] = i + 1

    def get_top_n(self, n=10):
        return self.capabilities[:n]

    def get_by_priority(self, priority):
        return [c for c in self.capabilities if c["priority"] == priority]

    def get_by_sheet(self, sheet):
        return [c for c in self.capabilities if c["sheet"] == sheet]

    def get_summary_stats(self):
        if not self.capabilities:
            return {}
        scores = [c["scores"]["final"] for c in self.capabilities]
        return {
            "total_capabilities": len(self.capabilities),
            "high_priority": len(self.get_by_priority("HIGH")),
            "medium_priority": len(self.get_by_priority("MEDIUM")),
            "low_priority": len(self.get_by_priority("LOW")),
            "avg_score": round(np.mean(scores), 2),
            "max_score": round(max(scores), 2),
            "min_score": round(min(scores), 2),
            "std_dev": round(np.std(scores), 2),
        }

    def to_dataframe(self):
        rows = []
        for c in self.capabilities:
            row = {
                "Rank": c["rank"],
                "Capability": c["name"],
                "Tier": c["tier"],
                "Category": c["sheet"],
                "Priority": c["priority"],
                "Final Score": c["scores"]["final"],
                "Customer Impact": c["scores"]["customer_impact"],
                "Business Impact": c["scores"]["business_impact"],
                "Cost to Implement": c["scores"]["cost_to_implement"],
                "Survey Demand": c["survey_demand"],
            }
            if "development_quarter" in c:
                row["Dev Quarter"] = c["development_quarter"]
            rows.append(row)
        return pd.DataFrame(rows)


def run_scoring_pipeline(survey_path, xlsx_path, config_path=None):
    engine = ScoringEngine(config_path)
    survey = SurveyAnalyzer(survey_path)
    scorer = CapabilityScorer(xlsx_path, survey, engine)
    capabilities = scorer.load_and_score()
    insights = survey.get_customer_insights()
    stats = scorer.get_summary_stats()
    return {
        "capabilities": capabilities,
        "survey_insights": insights,
        "summary_stats": stats,
        "scorer": scorer,
        "engine": engine,
    }


if __name__ == "__main__":
    import sys
    survey = sys.argv[1] if len(sys.argv) > 1 else "data/survey.csv"
    xlsx = sys.argv[2] if len(sys.argv) > 2 else "data/capabilities.xlsx"
    results = run_scoring_pipeline(survey, xlsx)
    print(json.dumps({
        "summary": results["summary_stats"],
        "top_10": [{"rank": c["rank"], "name": c["name"], "score": c["scores"]["final"], "priority": c["priority"]}
                   for c in results["capabilities"][:10]],
        "survey": {k: v for k, v in results["survey_insights"].items() if k != "capability_demand"},
    }, indent=2))
